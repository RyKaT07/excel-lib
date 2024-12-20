import openpyxl.utils
import openpyxl.utils.exceptions
import pandas as pd
import openpyxl
import io
import zipfile
import env

def file_to_io_stream(path):
    with open(path, "rb") as file:
        file_stream = io.BytesIO(file.read())
    return file_stream

class ExcelFile:
    def __init__(self, file_stream):
        self.worksheet_count = 0
        try:
            workbook = self._validate_excel_file(file_stream)
            self.non_cell_objects = self._check_for_non_cell_objects(workbook, file_stream)
        except Exception as e:
            print(f"Unecpected error occured during loading file into openpyxl: {e}")
        self.worksheet = pd.read_excel(file_stream)
    
    def _validate_excel_file(self, file_stream):
        try:
            file_stream.seek(0)
            workbook = openpyxl.load_workbook(file_stream)
            self.worksheet_count = len(workbook.sheetnames)
            return workbook
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError("It is not a valid excel file")
        
    def _check_for_non_cell_objects(self, openpyxl_workbook_instance, file_stream):
        workbook = openpyxl_workbook_instance
        non_cell_objects = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            if worksheet._images:
                for image in worksheet._images:
                    if image.anchor._from:
                        anchor = image.anchor._from
                        anchor_info = f"Image anchored at cell {anchor.col + 1}{anchor.row + 1} in sheet '{sheet_name}'."
                    else:
                        anchor_info = f"Image not anchored to any cell in sheet '{sheet_name}'."
                    non_cell_objects.append(anchor_info)

            if worksheet._charts:
                for chart in worksheet._charts:
                    non_cell_objects.append(f"Chart found in sheet '{sheet_name}'.")
        
        return non_cell_objects
    
    def _check_for_images_in_archive(self, file_stream):
        images_found = []

        file_stream_copy = io.BytesIO(file_stream.getvalue())

        with zipfile.ZipFile(file_stream_copy, 'r') as zip_ref:
            image_files = [file for file in zip_ref.namelist() if file.startswith("xl/media/")]
            if image_files:
                for image_file in image_files:
                    images_found.append(f"Image found: {image_file}")
            else:
                images_found.append("No images found in xl/media/ folder")

        return images_found

    def get_non_cell_objects_info(self):
        if self.non_cell_objects:
            return "\n".join(self.non_cell_objects)
        return "No non-cell objects detected."
    
    def get_sheet_names(self):
        return self.workbook.sheetnames
    
    """_summary_ generates library[section_name]=row_number
    """
    def _identify_sections(self):
        sections = {}
        current_section = None
        for row_index, value in enumerate(self.worksheet.iloc[:, 0]):
            if isinstance(value, str) and value.isupper():
                if current_section:
                    sections[current_section][1] = row_index
                current_section = value.strip()
                sections[current_section] = [row_index, None]

        if current_section:
            sections[current_section][1] = self.worksheet.iloc[:, 0].last_valid_index()
        return sections
    
    def create_template_structure(self):
        template_structure = {
            "takeover": {
                "global_data": {},
                "contact_person": None,
                "responsible_person": None
            },
            "stations": {}
        }

        # Identify sections
        sections = self._identify_sections()

        # Find the key corresponding to global_data
        takeover_divider_key = next((key for key in env.SECTION_STATION_TAKEOVER_DIVIDER if key in sections), None)

        if takeover_divider_key:
            global_data = {}
            for row_index, row in self.worksheet.iloc[:sections[takeover_divider_key][0], :2].iterrows():
                value, key = row
                if pd.notna(key) and pd.notna(value):
                    global_data[key.strip()] = row_index
            template_structure["takeover"]["global_data"] = global_data

        # Find the key corresponding to contact_person
        contact_person_key = next((key for key in env.SECTION_CONTACT_PERSON if key in sections), None)
        if contact_person_key:
            contact_person = {}
            for row_index, row in self.worksheet.iloc[sections[contact_person_key][0]:sections[contact_person_key][1], :2].iterrows():
                value, key = row
                if pd.notna(key) and pd.notna(value):
                    contact_person[key] = row_index
            template_structure["takeover"]["contact_person"] = contact_person

        # Find the key corresponding to responsible_person
        responsible_person_key = next((key for key in env.SECTION_RESPONSIBLE_PERSON if key in sections), None)
        if responsible_person_key:
            responsible_person = {}
            for row_index, row in self.worksheet.iloc[sections[responsible_person_key][0]:sections[responsible_person_key][1], :2].iterrows():
                value, key = row
                if pd.notna(key) and pd.notna(value):
                    responsible_person[key] = row_index
            template_structure["takeover"]["responsible_person"] = responsible_person

        # Station data (from "STACJA ŁADOWANIA - DANE" section and onwards)
        station_structure = {}
        for section, section_range in sections.items():
            station_data = {}
            for row_index in range(section_range[0], section_range[1]):
                key = self.worksheet.iat[row_index, 1]
                if pd.notna(key):
                    station_data[key] = row_index
            station_structure[section] = station_data
        template_structure["stations"]=station_structure

        return template_structure

        
    def get_template_for_this_file(self, template):
        if template.worksheet_count != 1:
            print("Too many worksheets. Only first one will be taken as template into account.")
        form = template.worksheet.iloc[:, :2]
        self.discarded_data_info = []
        worksheet = self.worksheet
        number_of_discarded_rows = worksheet.shape[0] - form.shape[0]
        if number_of_discarded_rows > 0:
            self.discarded_data_info.append(f"Discarded rows below the form. Number of discarded rows: {number_of_discarded_rows}")
            worksheet = worksheet.iloc[form.shape[0]]
        elif number_of_discarded_rows < 0:
            self.discarded_data_info.append(f"Too little rows. Bad form.")
        
        template_col = form.iloc[:, 1]
        form_col = worksheet.iloc[:, 1]
        matching_values = []

        for index, template_value in template_col.items():
            matching_row = next((i for i, value in form_col.items() if template_value == value), -1)
            if matching_row == -1:
                if form.iloc[index, 0] == worksheet.iloc[index, 0]:
                    matching_row = form.iloc[index, 0]
            matching_values.append([index, template_value, matching_row])
        
        self.comparison_template = pd.DataFrame(matching_values, columns=['Template Index', 'Value', 'Form Index'])
        return self.comparison_template.values.tolist()

    def retrive_stations(self):
        if self.discarded_data_info:
            pass
        else:
            self.discarded_data_info = []
        stations = []
        length = self.comparison_template.shape[0]
        for loop_index, (column_name, column_data) in enumerate(self.worksheet.iloc[:, 2:].items()):
            station_data = self.comparison_template.iloc[:, 1:].copy().values.tolist()
            for form_index, value, index in self.comparison_template.itertuples(index=False):
                if index is not -1:
                    station_data[form_index][1] = column_data.iat[index]

            nones = pd.isna([row[1] for row in station_data]).sum()

            if length - nones <= 3:
                self.discarded_data_info.append(f"Column {loop_index + 2} not taken int account. Too little data.")
            else:
                stations.append(station_data)

        return stations
    def _update_rows_in_structure(self, data_section):
        updated_section = {}
        for key, expected_row in data_section.items():
            actual_label = self.worksheet.iloc[expected_row, 1] if expected_row < len(self.worksheet) else None
            if pd.notna(actual_label) and actual_label.strip() == key:
                updated_section[key] = expected_row
            else:
                found_row = -1
                for row_index, value in self.worksheet.iloc[:, 1].items():
                    if pd.notna(value) and value.strip() == key:
                        found_row = row_index
                        break
                updated_section[key] = found_row
        return updated_section

    def compare_structure_with_file(self, template):
        updated_structure = {
            "takeover": {
                "global_data": {},
                "contact_person": None,
                "responsible_person": None
            },
            "stations": {}
        }

        for section_name, takeover_section in template["takeover"].items():
            updated_structure["takeover"][section_name] = self._update_rows_in_structure(takeover_section)

        for section_name, station_section in template["stations"].items():
            updated_structure["stations"][section_name] = self._update_rows_in_structure(station_section)

        return updated_structure


    def create_data_structure_from_template(self, template):

        data_structure = self.compare_structure_with_file(template)

        collected_takeover_structures = []

        for column in range(2, self.worksheet.shape[1]):
            # Take global data from column
            current_global_data = {
                key: self.worksheet.iloc[row, column] if row < len(self.worksheet) else None for key, row in data_structure["takeover"]["global_data"].items()
            }

            # Check whether there is a struct with this global data
            matching_group = None
            for group in collected_takeover_structures:
                if group["global_data"] == current_global_data:
                    matching_group = group
                    break

            # If there is none, create new one
            if not matching_group:
                matching_group = {
                    "global_data": current_global_data,
                    "contact_person": None,
                    "responsible_person": None,
                    "stations": []
                }
                collected_takeover_structures.append(matching_group)

            # Compare contact person
            current_contact_person = {
                key: self.worksheet.iloc[row, column] if row < len(self.worksheet) else None for key, row in data_structure["takeover"]["contact_person"].items()
            }
            if matching_group["contact_person"] is None:
                matching_group["contact_person"] = current_contact_person
            elif matching_group["contact_person"] != current_contact_person:
                matching_group["contact_person"] = "Dla każdej stacji inna"

            # Compare responsible person
            current_responsible_person = {
                key: self.worksheet.iloc[row, column] if row < len(self.worksheet) else None for key, row in data_structure["takeover"]["responsible_person"].items()
            }
            if matching_group["responsible_person"] is None:
                matching_group["responsible_person"] = current_responsible_person
            elif matching_group["responsible_person"] != current_responsible_person:
                matching_group["responsible_person"] = "Dla każdej stacji inna"

            # Add station data
            station_data = {
                section: {
                    field: self.worksheet.iloc[row, column] if row < len(self.worksheet) else None
                    for field, row in fields.items()
                }
                for section, fields in data_structure["stations"].items()
            }
            matching_group["stations"].append(station_data)

        return collected_takeover_structures

